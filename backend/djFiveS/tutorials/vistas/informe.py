import os
import openpyxl
from tutorials.models import Almacen, Compra, Venta
from datetime import datetime
from django.conf import settings
from django.shortcuts import render
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from reportlab.lib.units import cm
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle


#Principal
def modulo_informe (request):
    return render(request, 'informe.html')

#Lista Informe compra
def informecompra_lista(request):
    # Obtenemos todas las compras realizadas
    compras = Compra.objects.all()

    # Agregar lógica para filtrar por fecha si tienes el filtro en tu formulario
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        compras = compras.filter(fecha__range=[fecha_inicio, fecha_fin])
    
    context = {
        'compras': compras,
    }
    
    return render(request, 'informecompras_lista.html', context)

def informeventa_lista(request):
    # Obtenemos todas las compras realizadas
    ventas = Venta.objects.all()

    # Agregar lógica para filtrar por fecha si tienes el filtro en tu formulario
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        ventas = ventas.filter(fecha__range=[fecha_inicio, fecha_fin])
    
    context = {
        'ventas': ventas,
    }
    
    return render(request, 'informeventas_lista.html', context)

def informestock_lista(request):
    # Obtenemos todas las compras realizadas
    almacen = Almacen.objects.all()

    # Agregar lógica para filtrar por fecha si tienes el filtro en tu formulario
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    if fecha_inicio and fecha_fin:
        almacen = almacen.filter(fecha__range=[fecha_inicio, fecha_fin])
    
    context = {
        'almacen': almacen,
    }
    
    return render(request, 'informestock_lista.html', context)



#Compra Excel

def compra_excel(request):
    # Obtener los rangos de fechas desde la solicitud
    start_date = request.GET.get('fecha_inicio')
    end_date = request.GET.get('fecha_fin')

    # Analizar fechas si no son None
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Crear un libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compras"

    # Escribir encabezados de la tabla
    ws.append(["Nro", "Fecha", "Proveedor", "Producto", "Cantidad", "Total"])

    # Filtrar las compras por el rango de fechas si ambos valores son válidos
    compras = Compra.objects.all()
    if start_date and end_date:
        compras = compras.filter(fecha__range=[start_date, end_date])

    for idx, compra in enumerate(compras, 1):
        ws.append([idx, compra.fecha, compra.proveedor.nombre, compra.producto.nombre, compra.cantidad, compra.total])

    # Preparar la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="informe_compras.xlsx"'
    wb.save(response)

    return response

#Compra PDF
def compra_pdf(request):
    # Obtener los rangos de fechas desde la solicitud
    start_date = request.GET.get('fecha_inicio')
    end_date = request.GET.get('fecha_fin')

    # Analizar fechas si no son None
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Crear la respuesta HTTP con el tipo de contenido adecuado
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_compras.pdf"'

    # Crear el documento PDF con márgenes de 2 cm
    pdf = SimpleDocTemplate(response, pagesize=letter, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []

    # Estilos de texto
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=14, alignment=1, leading=14, fontName="Helvetica-Bold")
    normal_style = styles['BodyText']

    # Agregar logotipo centrado (logo más ancho)
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'FiveStore.jpg')
    logo = Image(logo_path, width=200, height=80)  # Logo más ancho
    logo.hAlign = 'CENTER'

    # Encabezado principal con cuadro gris más pequeño, más alto, más redondeado, y más separado del logo
    header_text = Table([
        [Paragraph("<b>Five Store Bolivia</b><br/>Año: 2024<br/>Dirección: av. Ejercito Nacional", normal_style)]
    ], colWidths=[280])  # Reduce el ancho del cuadro gris

    header_text.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('ROUNDRECT', (0, 0), (-1, -1), 20, 1, colors.black),
    ]))

    header_data = [
        [logo, header_text]
    ]
    header_table = Table(header_data, colWidths=[200, 300])  # Aumenta el ancho del logo y separa más el cuadro del logo
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.1 * cm))
    elements.append(Paragraph('Reporte de Compras', title_style))
    elements.append(Spacer(1, 0.2 * cm))

    # Línea de separación con sombra
    line = Table([[""]], colWidths=[480])
    line.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, colors.grey),
    ]))
    elements.append(line)
    elements.append(Spacer(1, 0.3 * cm))

    # Información de registro con la fecha actual
    current_date = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    registro_info_data = [
        [Paragraph(f"<b>Fecha Registro:</b> {current_date}", normal_style), Paragraph("<b>Usuario Registro:</b> ADMIN", normal_style)]
    ]
    registro_info_table = Table(registro_info_data, colWidths=[240, 240])
    registro_info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(registro_info_table)
    elements.append(Spacer(1, 0.3 * cm))

    # Filtrar las compras por fecha si ambos valores son válidos
    compras = Compra.objects.all()
    if start_date and end_date:
        compras = compras.filter(fecha__range=[start_date, end_date])

    # Detalles de compras y cálculo del total
    data = [["Fecha", "Producto", "Precio Compra", "Cantidad", "Sub Total"]]
    total_sum = 0
    for compra in compras:
        subtotal = compra.total  # Asumiendo que compra.total contiene el subtotal
        total_sum += subtotal
        data.append([compra.fecha.strftime('%Y-%m-%d'), compra.producto.nombre, compra.precioc, compra.cantidad, subtotal])

    # Tabla de detalles de productos con ancho total de 480 unidades
    detail_table = Table(data, colWidths=[80, 120, 80, 80, 120], hAlign='CENTER')
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Fondo blanco para las filas de datos
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(detail_table)
    elements.append(Spacer(1, 0.3 * cm))  # Separar la tabla de monto total

    # Total dinámico basado en el cálculo de total_sum
    total_data = [
        [Paragraph("<b>Monto Total</b>", normal_style), f"{total_sum:.2f}"]
    ]
    total_table = Table(total_data, colWidths=[400, 80])  # Ancho total de 480 unidades
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 0), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
    ]))
    elements.append(total_table)

    # Generar el PDF con los elementos creados
    pdf.build(elements)
    return response

def venta_excel(request):
    # Obtener los rangos de fechas desde la solicitud
    start_date = request.GET.get('fecha_inicio')
    end_date = request.GET.get('fecha_fin')

    # Analizar fechas si no son None
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Crear un libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"

    # Escribir encabezados de la tabla
    ws.append(["Nro", "Fecha", "Cliente", "Usuario Venta", "Producto", "Cantidad", "Precio Venta", "Total"])

    # Filtrar las ventas por el rango de fechas si ambos valores son válidos
    ventas = Venta.objects.all()
    if start_date and end_date:
        ventas = ventas.filter(fecha__range=[start_date, end_date])

    for idx, venta in enumerate(ventas, 1):
        # Obtener el primer detalle de la venta
        primer_detalle = venta.detalles.first()  # Asegúrate de usar el related_name definido en el modelo
        usuario_venta = primer_detalle.usuario.username if primer_detalle and hasattr(primer_detalle, 'usuario') else "N/A"

        # Iterar sobre los detalles de la venta para incluir todos los productos
        for detalle in venta.detalles.all():
            ws.append([
                idx,
                venta.fecha.strftime('%d/%m/%Y'),
                venta.cliente.nombre,
                usuario_venta,
                detalle.producto.nombre,
                detalle.cantidad,
                f"{detalle.preciov:.2f}",
                f"{detalle.total:.2f}"
            ])

    # Preparar la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="informe_ventas.xlsx"'
    wb.save(response)

    return response

#Venta PDF
def venta_pdf(request):
    # Obtener los rangos de fechas desde la solicitud
    start_date = request.GET.get('fecha_inicio')
    end_date = request.GET.get('fecha_fin')

    # Analizar fechas si no son None
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Crear la respuesta HTTP con el tipo de contenido adecuado
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_ventas.pdf"'

    # Crear el documento PDF con márgenes de 2 cm
    pdf = SimpleDocTemplate(response, pagesize=letter, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []

    # Estilos de texto
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=14, alignment=1, leading=14, fontName="Helvetica-Bold")
    normal_style = styles['BodyText']

    # Agregar logotipo centrado (logo más ancho)
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'FiveStore.jpg')
    logo = Image(logo_path, width=200, height=80)  # Logo más ancho
    logo.hAlign = 'CENTER'

    # Encabezado principal con cuadro gris redondeado
    header_text = Table([
        [Paragraph("<b>Five Store Bolivia</b><br/>Año: 2024<br/>Dirección: av. Ejercito Nacional", normal_style)]
    ], colWidths=[280])

    header_text.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('ROUNDRECT', (0, 0), (-1, -1), 20, 1, colors.black),
    ]))

    header_data = [
        [logo, header_text]
    ]
    header_table = Table(header_data, colWidths=[200, 300])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.3 * cm))

    # Título del reporte antes de la línea
    elements.append(Paragraph('Reporte de Ventas', title_style))
    elements.append(Spacer(1, 0.2 * cm))

    # Línea de separación con sombra
    line = Table([[""]], colWidths=[480])
    line.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, colors.grey),
    ]))
    elements.append(line)
    elements.append(Spacer(1, 0.3 * cm))

    # Información de registro con la fecha de generación del PDF
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    registro_info_data = [
        [Paragraph(f"<b>Fecha Registro:</b> {current_date}", normal_style), Paragraph("<b>Usuario Registro:</b> ADMIN", normal_style)]
    ]
    registro_info_table = Table(registro_info_data, colWidths=[240, 240])
    registro_info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(registro_info_table)
    elements.append(Spacer(1, 0.3 * cm))

    # Escribir encabezados de la tabla y filtrar las ventas por fecha si ambos valores son válidos
    data = [["Fecha", "Cliente", "Producto", "Cantidad", "Total"]]
    ventas = Venta.objects.all()
    if start_date and end_date:
        ventas = ventas.filter(fecha__range=[start_date, end_date])

    total_sum = 0  # Variable para el monto total
    for venta in ventas:
        data.append([venta.fecha.strftime('%Y-%m-%d'), venta.cliente.nombre, venta.producto.nombre, venta.cantidad, venta.total])
        total_sum += venta.total  # Sumar el valor de cada venta

    # Crear la tabla y estilizarla con fondo blanco en las celdas de datos
    table = Table(data, colWidths=[80, 120, 120, 80, 80], hAlign='CENTER')  # Ancho total = 480
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Encabezado gris
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Celdas de datos en blanco
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ])
    table.setStyle(style)
    elements.append(table)

    # Agregar el monto total
    total_data = [
        [Paragraph("<b>Monto Total</b>", normal_style), f"{total_sum:.2f}"]
    ]
    total_table = Table(total_data, colWidths=[400, 80])  # Ancho total = 480
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
    ]))
    elements.append(Spacer(1, 0.3 * cm))
    elements.append(total_table)

    # Generar el PDF con los elementos creados
    pdf.build(elements)
    return response

#Stock excel
def stock_excel(request):
    # Obtener los rangos de fechas desde la solicitud
    start_date = request.GET.get('fecha_inicio')
    end_date = request.GET.get('fecha_fin')

    # Analizar fechas si no son None
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Crear un libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock"

    # Escribir encabezados de la tabla
    ws.append(["Nro", "Fecha", "Producto", "Categoria", "Cantidad", "Precio Compra", "Precio Venta"])

    # Filtrar el stock por el rango de fechas si ambos valores son válidos
    almacenes = Almacen.objects.all()
    if start_date and end_date:
        almacenes = almacenes.filter(fecha__range=[start_date, end_date])

    for idx, almacen in enumerate(almacenes, 1):
        ws.append([idx, almacen.fecha, almacen.producto.nombre, almacen.categoria.nombre, almacen.cantidad, almacen.precioc, almacen.preciov])

    # Preparar la respuesta HTTP con el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="informe_stock.xlsx"'
    wb.save(response)

    return response

#Stock PDF
def stock_pdf(request):
    # Obtener los rangos de fechas desde la solicitud
    start_date = request.GET.get('fecha_inicio')
    end_date = request.GET.get('fecha_fin')

    # Analizar fechas si no son None
    if start_date:
        start_date = parse_date(start_date)
    if end_date:
        end_date = parse_date(end_date)

    # Crear la respuesta HTTP con el tipo de contenido adecuado
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="informe_stock.pdf"'

    # Crear el documento PDF con márgenes de 2 cm
    pdf = SimpleDocTemplate(response, pagesize=letter, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    elements = []

    # Estilos de texto
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(name='Title', fontSize=14, alignment=1, leading=14, fontName="Helvetica-Bold")
    normal_style = styles['BodyText']

    # Agregar logotipo centrado (logo más ancho)
    logo_path = os.path.join(settings.STATICFILES_DIRS[0], 'img', 'FiveStore.jpg')
    logo = Image(logo_path, width=200, height=80)  # Logo más ancho
    logo.hAlign = 'CENTER'

    # Encabezado principal con cuadro gris redondeado
    header_text = Table([
        [Paragraph("<b>Five Store Bolivia</b><br/>Año: 2024<br/>Dirección: av. Ejercito Nacional", normal_style)]
    ], colWidths=[280])

    header_text.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
        ('TOPPADDING', (0, 0), (-1, -1), 15),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 15),
        ('ROUNDRECT', (0, 0), (-1, -1), 20, 1, colors.black),
    ]))

    header_data = [
        [logo, header_text]
    ]
    header_table = Table(header_data, colWidths=[200, 300])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
    ]))

    elements.append(header_table)
    elements.append(Spacer(1, 0.3 * cm))

    # Título del reporte antes de la línea
    elements.append(Paragraph('Reporte de Stock', title_style))
    elements.append(Spacer(1, 0.2 * cm))

    # Línea de separación con sombra
    line = Table([[""]], colWidths=[480])
    line.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 2, colors.grey),
    ]))
    elements.append(line)
    elements.append(Spacer(1, 0.3 * cm))

    # Información de registro con la fecha de generación del PDF
    current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    registro_info_data = [
        [Paragraph(f"<b>Fecha Registro:</b> {current_date}", normal_style), Paragraph("<b>Usuario Registro:</b> ADMIN", normal_style)]
    ]
    registro_info_table = Table(registro_info_data, colWidths=[240, 240])
    registro_info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    elements.append(registro_info_table)
    elements.append(Spacer(1, 0.3 * cm))

    # Escribir encabezados de la tabla y filtrar el stock por fecha si ambos valores son válidos
    data = [["Producto", "Categoria", "Cantidad", "Precio Compra", "Precio Venta"]]
    almacenes = Almacen.objects.all()
    if start_date and end_date:
        almacenes = almacenes.filter(fecha__range=[start_date, end_date])

    for almacen in almacenes:
        data.append([almacen.producto.nombre, almacen.categoria.nombre, almacen.cantidad, almacen.precioc, almacen.preciov])

    # Crear la tabla y estilizarla con encabezado gris y fondo blanco para las filas
    table = Table(data, colWidths=[120, 120, 60, 80, 80], hAlign='CENTER')
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),  # Encabezado en gris
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),  # Fondo blanco para las filas
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ])
    table.setStyle(style)
    elements.append(table)

    # Generar el PDF con los elementos creados
    pdf.build(elements)
    return response